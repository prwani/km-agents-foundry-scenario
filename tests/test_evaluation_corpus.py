import json
import pathlib
import unittest
import zipfile

from openpyxl import load_workbook
from pptx import Presentation


ROOT = pathlib.Path(__file__).resolve().parents[1]
CORPUS = ROOT / "evaluation" / "corpus" / "v1"


class EvaluationCorpusTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.manifest = json.loads((CORPUS / "manifest.json").read_text(encoding="utf-8"))

    def test_manifest_defines_72_synthetic_runs(self):
        self.assertTrue(self.manifest["synthetic_only"])
        self.assertEqual(len(self.manifest["cases"]), 12)
        self.assertEqual(self.manifest["top_level_run_count"], 72)
        self.assertEqual(self.manifest["implementations"], ["prompt", "hosted"])

    def test_all_supported_formats_are_covered_and_readable(self):
        paths = {
            CORPUS / source["path"]
            for case in self.manifest["cases"]
            for source in case["sources"]
        }
        self.assertEqual({path.suffix for path in paths}, {".docx", ".pptx", ".pdf", ".xlsx"})
        for path in paths:
            self.assertTrue(path.is_file(), path)
            self.assertGreater(path.stat().st_size, 0, path)
            if path.suffix == ".docx":
                self.assertTrue(zipfile.is_zipfile(path), path)
            elif path.suffix == ".pptx":
                self.assertGreaterEqual(len(Presentation(path).slides), 1, path)
            elif path.suffix == ".xlsx":
                workbook = load_workbook(path, read_only=True, data_only=False)
                self.assertIn("Outcomes", workbook.sheetnames)
                workbook.close()
            elif path.suffix == ".pdf":
                self.assertEqual(path.read_bytes()[:5], b"%PDF-", path)

    def test_hard_gate_cases_are_explicit(self):
        by_id = {case["id"]: case for case in self.manifest["cases"]}
        self.assertTrue(by_id["case-06-sensitive"]["expected"]["must_fail_closed"])
        self.assertEqual(
            by_id["case-07-name-unapproved"]["expected"]["expected_customer_display_name"],
            "Customer",
        )
        self.assertIn("template-trap", by_id["case-05-template-trap"]["tags"])


if __name__ == "__main__":
    unittest.main()
